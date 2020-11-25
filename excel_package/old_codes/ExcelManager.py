from openpyxl import *
import sqlite3
from excel_package.old_codes.DataLoader import DataLoader
import pandas as pd


class ExcelManager:

    def __init__(self):

        # self.readDatabase()
        # self.updateSceneAttributes()
        self.dataLoader = DataLoader()
        self.conn = None

        wb = Workbook()
        # ws = wb.active
        # ws['A1'] = 230
        # ws['A2'] = datetime.datetime.now()
        # wb.save("Scenario_Data_Main.xlsx")

    def readDatabase(self):

        data = {}
        scn_states = []
        conn = sqlite3.connect('../AlgorithmStore.db')

        query = 'SELECT * FROM ScenarioState'
        states = conn.execute(query)

        for state in states:
            scn_states.append(state[1])

        query = 'SELECT * FROM Scenarios'
        scenarios = conn.execute(query)

        for scn in scenarios:
            scene = scn[1]
            data.update({scene : []})
            for state in scn_states:
                data[scene].append(state)

        conn.close()
        print(data)

    def updateSceneAttributes(self):

        attributes = ["NAME", "FIELD_SIZE", "FIELD_SIZE_CHANGE", "EVENT_STATE", "PRIMARY_FEATURE", "OBJECTIVE_ENERGY", "OBJECTIVE_BANDWIDTH", "OBJECTIVE_LATENCY", "OBJECTIVE_ACCURACY", "MIN_ENERGY_CONSUMPTION", "MAX_ENERGY_CONSUMPTION", "MIN_BANDWIDTH_CONSUMPTION", "MAX_BANDWIDTH_CONSUMPTION", "MIN_LATENCY", "MAX_LATENCY",  "SAMPLING_RATE", "SPREAD_RATE", "PRIMARY_MEDIUM", "HOMOGENOUS_NODES", "SIMUL_AGGREGATION_NODE_RATIO", "SIMUL_AGGREGATION_TYPE", "REPORTING_BY_PERIODIC", "REPORTING_BY_EVENT", "PERIODIC_MONITORING_BY_REALTIME", "PERIODIC_MONITORING_BY_QUERY", "PERIODIC_MONITORING_BY_EVENT", "LOCATION_AWARENESS", "NODE_MOBILITY", "BEST_TECHNIQUE"]
        conn = sqlite3.connect('../AlgorithmStore.db')

        for attribute in attributes:
            data = [attribute]
            query1 = "Select * from wsnAttribute where attribute_name like '%{}%'".format(attribute)
            query2 = "Insert into wsnAttribute(attribute_name) Values ('{}')".format(attribute)
            results = conn.execute(query1)
            data_exists = False
            for r in results:
                data_exists = True
                break

            if not data_exists:
                conn.execute(query2)

        conn.commit()
        conn.close()

    def createScenarioAttributeLink(self, scenario, attribute, _value, source=None):

        scn_app_id = 0

        scenario_id = self.getId(self.conn, 'Scenarios', 'app_name', scenario)
        attribute_id = self.getId(self.conn, 'wsnAttributes', 'attribute_name', attribute)
        attrib_value_id = self.getId(self.conn, 'wsnAttributeValues', 'value', _value)

        # queryCheck = "Select * from Scenarios Where scenario={}".format(scenario_id)
        # last_scn_app_id = self.if_exists(queryCheck)

        if attrib_value_id == 0:
            query = 'Insert into wsnAttributeValues(value) Values("{}")'.format(_value)
            cursor = self.conn.execute(query)
            attrib_value_id = cursor.lastrowid

        queryDelete = "Delete from ScenarioWsnAttributes Where scenario_application={} and wsnAttribute={} and " \
                     "wsnAttributeValue={}".format(last_scn_app_id, attribute_id, last_value_id)
        queryCheck = "Select * from ScenarioWsnAttributes Where scenario_application={} and wsnAttribute={} and " \
                     "wsnAttributeValue={}".format(last_scn_app_id, attribute_id, last_value_id)
        queryInsert = "Insert into ScenarioWsnAttributes(scenario_application, wsnAttribute, wsnAttributeValue) " \
                      "Values({}, {}, {})".format(last_scn_app_id, attribute_id, last_value_id)
        last_id = self.if_exists(queryDelete)
        if last_id <= 0:
            self.conn.execute(queryInsert)


    def getId(self, conn, table, field, value):

        id = 0
        query = 'Select * from {} Where lower({}) like lower("{}")'.format(table, field, value)
        results = conn.execute(query)

        for row in results:
            id = row[0]
            break

        return id

    def if_exists(self, query, delete=False, deleteQuery=None):

        last_id = 0
        data_exists = False
        cursor = self.conn.execute(query)
        for c in cursor:
            last_id = c[0]
            break

        return last_id


    def loadData(self):

        self.conn = sqlite3.connect('../AlgorithmStore.db')

        self.dataLoader.loadEarthquake_Detection(self.createScenarioAttributeLink)

        self.conn.commit()
        self.conn.close()

        # self.printData()
        self.writeToFile()
        print('All Processed...')

    def printData(self):

        data = []
        conn = sqlite3.connect('../AlgorithmStore.db')
        query = 'Select Scenarios.app_name, ScenarioState.state_name, wsnAttributes.attribute_name, wsnAttributeValues.value from Scenarios, ScenarioState, wsnAttributes, wsnAttributeValues, ScenarioApplications, ScenarioWsnAttributes where Scenarios.id=ScenarioApplications.scenario and ScenarioApplications.id=ScenarioWsnAttributes.scenario_application and wsnAttributes.id=ScenarioWsnAttributes.wsnAttribute and wsnAttributeValues.id=ScenarioWsnAttributes.wsnAttributeValue and ScenarioState.id=ScenarioApplications.event_state order by app_name, state_name, attribute_name'
        cursor = conn.execute(query)
        for c in cursor:
            data.append({'app_name': c[0], 'state_name': c[1], 'attribute': c[2], 'value': c[3]})
        conn.close()

        dataDF = pd.DataFrame(data)

    def writeToFile(self):

        filename = '../wsn_data.csv'
        conn = sqlite3.connect('../AlgorithmStore.db')
        app_id = 1
        query = 'Select Scenarios.app_name, ScenarioState.state_name, wsnAttributes.attribute_name, wsnAttributeValues.value from Scenarios, ScenarioState, wsnAttributes, wsnAttributeValues, ScenarioApplications, ScenarioWsnAttributes where Scenarios.id=ScenarioApplications.scenario and ScenarioApplications.id=ScenarioWsnAttributes.scenario_application and wsnAttributes.id=ScenarioWsnAttributes.wsnAttribute and wsnAttributeValues.id=ScenarioWsnAttributes.wsnAttributeValue and ScenarioState.id=ScenarioApplications.event_state and ScenarioApplications.id={} order by app_name, state_name, attribute_name'.format(app_id)
        cursor = conn.execute(query)

        f_titles = []
        f_data = []
        for c in cursor:
            latest_data = {'app_name': c[0], 'state_name': c[1], 'attribute': c[2], 'value': c[3]}

            f_titles. append(latest_data['attribute'])
            f_data.append('value')

        conn.close()
        data = pd.DataFrame(f_data, columns=f_titles)
        print(data)
